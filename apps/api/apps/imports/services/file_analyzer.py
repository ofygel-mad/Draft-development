import logging
import os
from tempfile import NamedTemporaryFile
from typing import Any

from django.core.files.storage import default_storage

logger = logging.getLogger(__name__)

# Auto-mapping hints: (keywords → crm_field)
FIELD_HINTS = {
    'full_name': ['имя', 'name', 'фио', 'клиент', 'ф.и.о', 'contact'],
    'phone': ['телефон', 'phone', 'тел', 'mobile', 'номер'],
    'email': ['email', 'почта', 'e-mail', 'mail'],
    'company_name': ['компания', 'company', 'организация', 'фирма'],
    'source': ['источник', 'source', 'откуда', 'канал'],
    'status': ['статус', 'status', 'состояние'],
}


def auto_detect_mapping(headers: list[str]) -> dict[str, str]:
    """Возвращает {excel_column: crm_field} для автоматического маппинга."""
    mapping = {}
    for header in headers:
        h_lower = header.lower().strip()
        for crm_field, hints in FIELD_HINTS.items():
            if any(hint in h_lower for hint in hints):
                mapping[header] = crm_field
                break
    return mapping


def analyze_file(file_path: str, import_type: str) -> dict[str, Any]:
    """Читает файл и возвращает preview + auto-mapping."""
    ext = os.path.splitext(file_path)[1].lower()

    try:
        if ext in ('.xlsx', '.xls'):
            return _analyze_excel(file_path)
        if ext == '.csv':
            return _analyze_csv(file_path)
        raise ValueError(f'Unsupported file type: {ext}')
    except Exception as exc:
        logger.exception('File analysis failed: %s', exc)
        raise


def _analyze_excel(file_path: str) -> dict:
    import openpyxl

    if default_storage.exists(file_path):
        with default_storage.open(file_path, 'rb') as source, NamedTemporaryFile(suffix=os.path.splitext(file_path)[1].lower()) as tmp:
            tmp.write(source.read())
            tmp.flush()
            wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
    else:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

    if not rows:
        return {'headers': [], 'rows': [], 'total': 0, 'auto_mapping': {}}

    headers = [str(h) if h is not None else '' for h in rows[0]]
    preview_rows = [
        [str(cell) if cell is not None else '' for cell in row]
        for row in rows[1:21]  # first 20 rows
    ]
    total = len(rows) - 1

    return {
        'headers': headers,
        'rows': preview_rows,
        'total': total,
        'auto_mapping': auto_detect_mapping(headers),
    }


def _analyze_csv(file_path: str) -> dict:
    import csv

    if default_storage.exists(file_path):
        with default_storage.open(file_path, 'r', encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            rows = list(reader)
    else:
        with open(file_path, encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            rows = list(reader)

    if not rows:
        return {'headers': [], 'rows': [], 'total': 0, 'auto_mapping': {}}

    headers = rows[0]
    preview_rows = rows[1:21]
    total = len(rows) - 1

    return {
        'headers': headers,
        'rows': preview_rows,
        'total': total,
        'auto_mapping': auto_detect_mapping(headers),
    }
