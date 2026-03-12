import csv
import logging
import re
from tempfile import NamedTemporaryFile
from typing import Any, Iterator

import phonenumbers
from django.core.files.storage import default_storage
from django.db import IntegrityError

logger = logging.getLogger(__name__)

BATCH_SIZE = 500


def _normalize_phone(phone: str) -> str:
    if not phone:
        return phone
    try:
        parsed = phonenumbers.parse(phone, 'KZ')
        if phonenumbers.is_valid_number(parsed):
            return phonenumbers.format_number(parsed, phonenumbers.PhoneNumberFormat.E164)
    except phonenumbers.NumberParseException:
        logger.debug('Could not parse phone during import: %s', phone)
    return re.sub(r'[\s\-\(\)\+]', '', phone)


class ImportProcessor:
    def __init__(self, job):
        self.job = job

    def run(self) -> dict[str, Any]:
        mapping = self.job.column_mapping
        if not mapping:
            raise ValueError('Column mapping is empty')

        if self.job.import_type == 'customer':
            return self._import_customers(self._iter_rows(), mapping)
        raise ValueError(f'Unknown import_type: {self.job.import_type}')

    def _iter_rows(self) -> Iterator[list[str]]:
        ext = self.job.file_name.rsplit('.', 1)[-1].lower()
        if ext == 'csv':
            yield from self._iter_csv_rows()
            return
        if ext in ('xlsx', 'xls'):
            yield from self._iter_excel_rows()
            return
        raise ValueError(f'Unsupported format: {ext}')

    def _iter_csv_rows(self) -> Iterator[list[str]]:
        if default_storage.exists(self.job.file_path):
            with default_storage.open(self.job.file_path, 'r', encoding='utf-8-sig', newline='') as f:
                reader = csv.reader(f)
                next(reader, None)  # header
                for row in reader:
                    yield [str(cell) if cell is not None else '' for cell in row]
            return

        with open(self.job.file_path, encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            next(reader, None)  # header
            for row in reader:
                yield [str(cell) if cell is not None else '' for cell in row]

    def _iter_excel_rows(self) -> Iterator[list[str]]:
        import openpyxl

        if default_storage.exists(self.job.file_path):
            with default_storage.open(self.job.file_path, 'rb') as source, NamedTemporaryFile(suffix='.xlsx') as tmp:
                tmp.write(source.read())
                tmp.flush()
                wb = openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
                try:
                    ws = wb.active
                    iterator = ws.iter_rows(values_only=True)
                    next(iterator, None)  # header
                    for row in iterator:
                        yield [str(c) if c is not None else '' for c in row]
                finally:
                    wb.close()
            return

        wb = openpyxl.load_workbook(self.job.file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            iterator = ws.iter_rows(values_only=True)
            next(iterator, None)  # header
            for row in iterator:
                yield [str(c) if c is not None else '' for c in row]
        finally:
            wb.close()

    def _import_customers(self, rows: Iterator[list[str]], mapping: dict) -> dict[str, Any]:
        from apps.customers.models import Customer

        created = updated = 0
        errors = []

        preview = self.job.preview_json
        headers = preview.get('headers', [])
        index_mapping: dict[int, str] = {}
        for key, field in mapping.items():
            try:
                index_mapping[int(key)] = field
            except ValueError:
                if key in headers:
                    index_mapping[headers.index(key)] = field

        batch_payload: list[tuple[int, dict[str, str]]] = []

        def flush_batch(payload: list[tuple[int, dict[str, str]]]):
            nonlocal created, updated
            if not payload:
                return

            phones = {item[1]['phone'] for item in payload if item[1].get('phone')}
            emails = {item[1]['email'] for item in payload if item[1].get('email')}

            existing_by_phone = {
                c.phone: c
                for c in Customer.objects.filter(
                    organization=self.job.organization,
                    phone__in=phones,
                    deleted_at__isnull=True,
                )
            } if phones else {}

            existing_by_email = {
                c.email: c
                for c in Customer.objects.filter(
                    organization=self.job.organization,
                    email__in=emails,
                    deleted_at__isnull=True,
                )
            } if emails else {}

            to_create: list[Customer] = []
            to_update: list[Customer] = []

            for row_num, data in payload:
                lookup_customer = existing_by_phone.get(data.get('phone', ''))
                if not lookup_customer and data.get('email'):
                    lookup_customer = existing_by_email.get(data['email'])

                if lookup_customer:
                    changed = False
                    for field, value in {
                        'full_name': data.get('full_name', 'Без имени'),
                        'email': data.get('email', ''),
                        'company_name': data.get('company_name', ''),
                        'source': data.get('source', 'import'),
                        'owner': self.job.created_by,
                    }.items():
                        if getattr(lookup_customer, field) != value:
                            setattr(lookup_customer, field, value)
                            changed = True
                    if changed:
                        to_update.append(lookup_customer)
                        updated += 1
                    continue

                to_create.append(
                    Customer(
                        organization=self.job.organization,
                        owner=self.job.created_by,
                        full_name=data.get('full_name', 'Без имени'),
                        phone=data.get('phone', ''),
                        email=data.get('email', ''),
                        company_name=data.get('company_name', ''),
                        source=data.get('source', 'import'),
                    )
                )
                created += 1

            if to_create:
                try:
                    Customer.objects.bulk_create(to_create, batch_size=BATCH_SIZE)
                except IntegrityError as exc:
                    logger.warning('Bulk create integrity error for job %s: %s', self.job.id, exc)
                    for customer in to_create:
                        try:
                            customer.save()
                        except IntegrityError as row_exc:
                            errors.append({'row': None, 'error': str(row_exc)})

            if to_update:
                Customer.objects.bulk_update(
                    to_update,
                    fields=['full_name', 'email', 'company_name', 'source', 'owner', 'updated_at'],
                    batch_size=BATCH_SIZE,
                )

        for row_idx, row in enumerate(rows, start=2):
            try:
                data: dict[str, str] = {}
                for col_idx, field in index_mapping.items():
                    if col_idx < len(row):
                        val = str(row[col_idx]).strip()
                        if val and val != 'None':
                            data[field] = val

                if data.get('phone'):
                    data['phone'] = _normalize_phone(data['phone'])

                if not data.get('full_name') and not data.get('phone'):
                    errors.append({'row': row_idx, 'error': 'Нет имени и телефона'})
                    continue

                batch_payload.append((row_idx, data))
                if len(batch_payload) >= BATCH_SIZE:
                    flush_batch(batch_payload)
                    batch_payload = []
            except Exception as exc:
                logger.warning('Import row %d error: %s', row_idx, exc)
                errors.append({'row': row_idx, 'error': str(exc)})

        flush_batch(batch_payload)

        return {'created': created, 'updated': updated, 'errors': errors[:100]}
