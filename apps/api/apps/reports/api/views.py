import io
from datetime import timedelta

from django.core.cache import cache
from django.db import models
from django.db.models import Count, Sum
from django.db.models.functions import TruncMonth
from django.http import HttpResponse, StreamingHttpResponse
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView


class DashboardSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        cache_key = f'dashboard:{request.user.organization_id}'
        cached = cache.get(cache_key)
        if cached:
            return Response(cached)

        from apps.customers.models import Customer
        from apps.deals.models import Deal
        from apps.tasks.models import Task
        from apps.activities.models import Activity

        org         = request.user.organization
        now         = timezone.now()
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        prev_start  = (month_start - timedelta(days=1)).replace(day=1)

        customers_total      = Customer.objects.filter(organization=org, deleted_at__isnull=True).count()
        customers_this_month = Customer.objects.filter(organization=org, created_at__gte=month_start).count()
        customers_prev_month = Customer.objects.filter(organization=org, created_at__gte=prev_start, created_at__lt=month_start).count()
        active_deals         = Deal.objects.filter(organization=org, status='open', deleted_at__isnull=True).count()
        revenue              = Deal.objects.filter(organization=org, status='won', closed_at__gte=month_start).aggregate(t=Sum('amount'))['t'] or 0
        tasks_today          = Task.objects.filter(organization=org, assigned_to=request.user, status='open', due_at__date=now.date()).count()
        overdue_tasks        = Task.objects.filter(organization=org, assigned_to=request.user, status='open', due_at__lt=now).count()

        three_days_ago = now - timedelta(days=3)
        active_deal_ids = list(Deal.objects.filter(organization=org, status='open', deleted_at__isnull=True).values_list('id', flat=True))
        deals_with_recent = Activity.objects.filter(
            organization=org,
            deal_id__in=active_deal_ids,
            created_at__gte=three_days_ago,
        ).values_list('deal_id', flat=True).distinct().count()
        deals_no_activity = max(len(active_deal_ids) - deals_with_recent, 0)

        five_days_ago = now - timedelta(days=5)
        stalled_deals = list(
            Deal.objects.filter(
                organization=org,
                status='open',
                deleted_at__isnull=True,
            ).filter(
                models.Q(last_activity_at__lt=five_days_ago)
                | models.Q(last_activity_at__isnull=True, created_at__lt=five_days_ago)
            ).select_related('customer', 'stage')
            .order_by('last_activity_at')[:5]
            .values(
                'id',
                'title',
                'amount',
                'currency',
                'last_activity_at',
                'stage__name',
                'customer__full_name',
                'customer__id',
            )
        )

        seven_days_ago = now - timedelta(days=7)
        silent_customers = list(
            org.customers.filter(
                deleted_at__isnull=True,
                deals__status='open',
            ).filter(
                models.Q(last_contact_at__lt=seven_days_ago)
                | models.Q(last_contact_at__isnull=True, created_at__lt=seven_days_ago)
            ).distinct()
            .order_by('last_contact_at')[:5]
            .values('id', 'full_name', 'company_name', 'last_contact_at', 'phone')
        )

        today_tasks_qs = Task.objects.filter(
            organization=org,
            assigned_to=request.user,
            status='open',
            due_at__date=now.date(),
        ).select_related('customer').order_by('due_at')[:5]
        today_tasks_list = [
            {
                'id': str(t.id),
                'title': t.title,
                'priority': t.priority,
                'due_at': t.due_at.isoformat() if t.due_at else None,
                'customer': {'id': str(t.customer_id), 'full_name': t.customer.full_name} if t.customer else None,
            }
            for t in today_tasks_qs
        ]

        deals_by_stage = list(Deal.objects.filter(organization=org, status='open').values('stage__name').annotate(count=Count('id'), amount=Sum('amount')).order_by('-amount')[:10])
        customers_by_source = list(Customer.objects.filter(organization=org, deleted_at__isnull=True).values('source').annotate(count=Count('id')).order_by('-count')[:8])
        revenue_by_month = list(Deal.objects.filter(organization=org, status='won', closed_at__isnull=False).annotate(month=TruncMonth('closed_at')).values('month').annotate(revenue=Sum('amount'), deals=Count('id')).order_by('month'))
        manager_stats = list(Deal.objects.filter(organization=org, status='won').values('owner__full_name').annotate(deals=Count('id'), revenue=Sum('amount')).order_by('-revenue')[:10])

        all_deals  = Deal.objects.filter(organization=org, deleted_at__isnull=True).count() or 1
        with_deals = Customer.objects.filter(organization=org, deals__isnull=False, deleted_at__isnull=True).distinct().count()
        won_deals  = Deal.objects.filter(organization=org, status='won').count()

        data = {
            'customers_count':   customers_total,
            'customers_delta':   customers_this_month - customers_prev_month,
            'active_deals_count': active_deals,
            'revenue_month':     float(revenue),
            'revenue_delta':     0,
            'tasks_today':       tasks_today,
            'overdue_tasks':     overdue_tasks,
            'deals_no_activity': deals_no_activity,
            'stalled_deals': [
                {
                    'id': str(d['id']),
                    'title': d['title'],
                    'amount': float(d['amount'] or 0),
                    'currency': d['currency'],
                    'stage': d['stage__name'],
                    'customer_name': d['customer__full_name'] or '',
                    'customer_id': str(d['customer__id'] or ''),
                    'days_silent': (now - d['last_activity_at']).days if d['last_activity_at'] else None,
                }
                for d in stalled_deals
            ],
            'silent_customers': [
                {
                    'id': str(c['id']),
                    'full_name': c['full_name'],
                    'company_name': c['company_name'] or '',
                    'phone': c['phone'] or '',
                    'days_silent': (now - c['last_contact_at']).days if c['last_contact_at'] else None,
                }
                for c in silent_customers
            ],
            'today_tasks':       today_tasks_list,
            'recent_customers':  list(Customer.objects.filter(organization=org, deleted_at__isnull=True).order_by('-created_at').values('id', 'full_name', 'company_name', 'status', 'created_at')[:5]),
            'deals_by_stage':    [{'stage': d['stage__name'] or 'Без этапа', 'count': d['count'], 'amount': float(d['amount'] or 0)} for d in deals_by_stage],
            'customers_by_source': [{'source': c['source'] or 'Не указан', 'count': c['count']} for c in customers_by_source],
            'revenue_by_month':  [{'month': r['month'].strftime('%b %Y'), 'revenue': float(r['revenue'] or 0), 'deals': r['deals']} for r in revenue_by_month],
            'manager_leaderboard': [{'name': m['owner__full_name'] or 'Неизвестно', 'deals': m['deals'], 'revenue': float(m['revenue'] or 0)} for m in manager_stats],
            'funnel': {
                'customers': customers_total, 'with_deals': with_deals, 'deals': all_deals, 'won': won_deals,
                'conversion_rate': round(won_deals / all_deals * 100, 1),
            },
        }
        cache.set(cache_key, data, timeout=300)
        return Response(data)


def _excel_response(wb, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def _header_style():
    from openpyxl.styles import Font, PatternFill, Alignment
    return (
        Font(bold=True, color='FFFFFF', size=11, name='Calibri'),
        PatternFill(start_color='D97706', end_color='D97706', fill_type='solid'),
        Alignment(horizontal='center', vertical='center'),
    )


class ExportCustomersExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import PatternFill, Alignment, Border, Side
        from apps.customers.models import Customer

        qs = Customer.objects.filter(organization=request.user.organization, deleted_at__isnull=True).select_related('owner').order_by('-created_at')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Клиенты'

        font, fill, align = _header_style()
        alt   = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
        bord  = Border(bottom=Side(style='thin', color='E5E7EB'))
        hdrs  = ['Имя', 'Компания', 'Телефон', 'Email', 'Статус', 'Источник', 'Ответственный', 'Дата']

        for col, h in enumerate(hdrs, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.alignment = font, fill, align
        ws.row_dimensions[1].height = 28

        st = {'new': 'Новый', 'active': 'Активный', 'inactive': 'Неактивный', 'archived': 'Архив'}
        for ri, obj in enumerate(qs, 2):
            for ci, val in enumerate([obj.full_name, obj.company_name or '', obj.phone or '', obj.email or '',
                st.get(obj.status, obj.status), obj.source or '',
                obj.owner.full_name if obj.owner else '', obj.created_at.strftime('%d.%m.%Y')], 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = bord
                cell.alignment = Alignment(vertical='center')
                if ri % 2 == 0:
                    cell.fill = alt
            ws.row_dimensions[ri].height = 20

        for col, w in enumerate([30, 25, 18, 30, 14, 20, 25, 16], 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
        ws.freeze_panes = 'A2'
        return _excel_response(wb, 'customers.xlsx')


class ExportDealsExcelView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import PatternFill, Alignment
        from apps.deals.models import Deal

        qs = Deal.objects.filter(organization=request.user.organization, deleted_at__isnull=True).select_related('customer', 'stage', 'pipeline', 'owner').order_by('-created_at')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Сделки'

        font, fill, align = _header_style()
        alt  = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
        hdrs = ['Название', 'Клиент', 'Воронка', 'Этап', 'Сумма', 'Валюта', 'Статус', 'Ответственный', 'Создана', 'Закрыта']

        for col, h in enumerate(hdrs, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.alignment = font, fill, align
        ws.row_dimensions[1].height = 28

        st = {'open': 'В работе', 'won': 'Выиграна', 'lost': 'Проиграна'}
        for ri, d in enumerate(qs, 2):
            row = [d.title, d.customer.full_name if d.customer else '', d.pipeline.name, d.stage.name,
                float(d.amount) if d.amount else '', d.currency, st.get(d.status, d.status),
                d.owner.full_name if d.owner else '', d.created_at.strftime('%d.%m.%Y'),
                d.closed_at.strftime('%d.%m.%Y') if d.closed_at else '']
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(vertical='center')
                if ri % 2 == 0:
                    cell.fill = alt
                if ci == 5 and val:
                    cell.number_format = '#,##0.00'
            ws.row_dimensions[ri].height = 20

        for col, w in enumerate([30, 25, 20, 20, 14, 10, 14, 25, 14, 14], 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = w
        ws.freeze_panes = 'A2'
        return _excel_response(wb, 'deals.xlsx')


class ReportExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.customers.models import Customer
        qs = Customer.objects.filter(organization=request.user.organization, deleted_at__isnull=True).values('full_name', 'company_name', 'phone', 'email', 'status', 'source', 'created_at')

        def stream():
            yield 'Имя,Компания,Телефон,Email,Статус,Источник,Дата\n'
            for c in qs:
                yield f"{c['full_name']},{c['company_name'] or ''},{c['phone'] or ''},{c['email'] or ''},{c['status']},{c['source'] or ''},{c['created_at'].strftime('%d.%m.%Y')}\n"

        resp = StreamingHttpResponse(stream(), content_type='text/csv; charset=utf-8')
        resp['Content-Disposition'] = 'attachment; filename="crm-export.csv"'
        return resp


class ManagerKpiView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.deals.models import Deal
        from apps.tasks.models import Task
        from apps.customers.models import Customer

        org = request.user.organization
        now = timezone.now()
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        managers = list(
            Deal.objects.filter(organization=org, created_at__gte=start)
            .values('owner__id', 'owner__full_name')
            .annotate(
                deals_open=Count('id', filter=models.Q(status='open')),
                deals_won=Count('id', filter=models.Q(status='won')),
                deals_lost=Count('id', filter=models.Q(status='lost')),
                revenue=Sum('amount', filter=models.Q(status='won')),
            )
            .order_by('-revenue')
        )

        result = []
        for m in managers:
            uid = m['owner__id']
            total = (m['deals_won'] or 0) + (m['deals_lost'] or 0)
            winrate = round(m['deals_won'] / total * 100, 1) if total else 0
            tasks_done = Task.objects.filter(
                organization=org,
                assigned_to_id=uid,
                status='done',
                updated_at__gte=start,
            ).count()
            new_customers = Customer.objects.filter(
                organization=org,
                owner_id=uid,
                created_at__gte=start,
                deleted_at__isnull=True,
            ).count()
            result.append({
                'id': str(uid),
                'name': m['owner__full_name'] or 'Неизвестно',
                'deals_open': m['deals_open'],
                'deals_won': m['deals_won'],
                'deals_lost': m['deals_lost'],
                'win_rate': winrate,
                'revenue': float(m['revenue'] or 0),
                'tasks_done': tasks_done,
                'new_customers': new_customers,
            })

        return Response({'managers': result, 'period': start.strftime('%B %Y')})
