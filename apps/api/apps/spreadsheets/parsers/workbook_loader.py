from pathlib import Path
from io import BytesIO

from django.core.files.storage import default_storage
from openpyxl import load_workbook


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def load_workbook_from_path(file_path: str):
    path = Path(file_path)
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported workbook extension: {path.suffix}")
    return load_workbook(filename=file_path, data_only=False, read_only=False)


def load_workbook_from_storage(storage_key: str):
    path = Path(storage_key)
    if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported workbook extension: {path.suffix}")

    if default_storage.exists(storage_key):
        with default_storage.open(storage_key, 'rb') as source:
            content = source.read()
        return load_workbook(filename=BytesIO(content), data_only=False, read_only=False)

    return load_workbook_from_path(storage_key)
