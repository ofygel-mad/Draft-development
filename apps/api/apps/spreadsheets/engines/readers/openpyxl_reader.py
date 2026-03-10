from openpyxl import load_workbook


def read_workbook_metadata(file_path: str) -> list[dict]:
    workbook = load_workbook(filename=file_path, data_only=False, read_only=True)
    sheets = []
    for position, sheet_name in enumerate(workbook.sheetnames):
        sheet = workbook[sheet_name]
        sheets.append(
            {
                "name": sheet_name,
                "position": position,
                "max_row": sheet.max_row,
                "max_col": sheet.max_column,
            }
        )
    workbook.close()
    return sheets
