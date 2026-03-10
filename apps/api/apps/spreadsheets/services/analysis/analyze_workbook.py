from openpyxl.utils import get_column_letter

from apps.spreadsheets.models import SpreadsheetDocument, SpreadsheetSheet, SpreadsheetStyleSnapshot, SpreadsheetVersion
from apps.spreadsheets.services.analysis.detect_formulas import detect_formulas
from apps.spreadsheets.services.analysis.detect_headers import detect_headers
from apps.spreadsheets.services.analysis.detect_merged_ranges import detect_merged_ranges
from apps.spreadsheets.services.analysis.detect_sheets import detect_sheets
from apps.spreadsheets.services.analysis.detect_styles import detect_style_stats


def analyze_workbook(*, version: SpreadsheetVersion, workbook) -> None:
    sheets = detect_sheets(workbook)
    headers = detect_headers(workbook)
    formulas = detect_formulas(workbook)
    merged_ranges = detect_merged_ranges(workbook)
    style_stats = detect_style_stats(workbook)

    SpreadsheetSheet.objects.filter(version=version).delete()
    SpreadsheetStyleSnapshot.objects.filter(version=version).delete()

    for sheet_meta in sheets:
        sheet_name = sheet_meta["name"]
        SpreadsheetSheet.objects.create(
            version=version,
            name=sheet_name,
            position=sheet_meta["position"],
            max_row=sheet_meta["max_row"],
            max_col=sheet_meta["max_col"],
            detected_table_ranges=[],
            metadata={
                "headers": headers.get(sheet_name, {}),
                "formula_count": formulas.get(sheet_name, 0),
                "style_stats": style_stats.get(sheet_name, {}),
            },
        )

        SpreadsheetStyleSnapshot.objects.create(
            version=version,
            sheet_name=sheet_name,
            range_ref=(
                f"A1:{get_column_letter(max(workbook[sheet_name].max_column, 1))}{max(workbook[sheet_name].max_row, 1)}"
            ),
            merged_ranges=merged_ranges.get(sheet_name, []),
            style=style_stats.get(sheet_name, {}),
            freeze_panes={"value": style_stats.get(sheet_name, {}).get("freeze_panes")},
            filters={"enabled": style_stats.get(sheet_name, {}).get("has_auto_filter")},
        )

    SpreadsheetDocument.objects.filter(id=version.document_id).update(
        status=SpreadsheetDocument.Status.READY,
        current_version=version,
    )
