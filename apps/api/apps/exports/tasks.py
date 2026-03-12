import logging
import csv
from pathlib import Path
from tempfile import NamedTemporaryFile

from celery import shared_task
from django.core.files.storage import default_storage
from django.utils import timezone
from openpyxl import load_workbook

from apps.audit.models import AuditLog
from apps.audit.services import log_action
from apps.spreadsheets.domain import SpreadsheetJobStatus
from apps.spreadsheets.models import SpreadsheetExportJob

logger = logging.getLogger(__name__)


def _ensure_parent_dir(storage_path: str) -> None:
    """Create parent directories when using filesystem-backed storage."""
    try:
        resolved_path = Path(default_storage.path(storage_path))
    except (AttributeError, NotImplementedError):
        return
    resolved_path.parent.mkdir(parents=True, exist_ok=True)


@shared_task(queue='exports', bind=True, max_retries=3)
def process_export(self, export_job_id: str) -> None:
    try:
        job = SpreadsheetExportJob.objects.select_related('document', 'version').get(id=export_job_id)
    except SpreadsheetExportJob.DoesNotExist:
        logger.warning('process_export: job not found export_job_id=%s', export_job_id)
        return

    if job.status in {SpreadsheetJobStatus.COMPLETED, SpreadsheetJobStatus.FAILED}:
        logger.info('process_export: skip terminal job export_job_id=%s status=%s', export_job_id, job.status)
        return

    now = timezone.now()
    SpreadsheetExportJob.objects.filter(id=job.id).update(
        status=SpreadsheetJobStatus.RUNNING,
        started_at=job.started_at or now,
        error_text='',
    )

    try:
        source_key = (job.version.storage_key if job.version_id else job.document.storage_key).strip()
        if not default_storage.exists(source_key):
            raise FileNotFoundError(f'Source file not found: {source_key}')

        source_suffix = source_key.rsplit('.', 1)[-1].lower() if '.' in source_key else 'xlsx'
        export_format = str(job.summary_json.get('format') or source_suffix or 'xlsx').lower()
        if export_format not in {'csv', 'xlsx'}:
            raise ValueError(f'Unsupported export format: {export_format}')

        out_rel = f'exports/{job.organization_id}/{job.id}.{export_format}'
        _ensure_parent_dir(out_rel)

        if export_format == 'xlsx':
            with default_storage.open(source_key, 'rb') as src, default_storage.open(out_rel, 'wb') as dst:
                dst.write(src.read())
        else:
            if source_suffix == 'csv':
                with default_storage.open(source_key, 'rb') as src, default_storage.open(out_rel, 'wb') as dst:
                    dst.write(src.read())
            else:
                with default_storage.open(source_key, 'rb') as source, NamedTemporaryFile(suffix='.xlsx') as tmp:
                    tmp.write(source.read())
                    tmp.flush()
                    workbook = load_workbook(filename=tmp.name, data_only=True, read_only=True)
                    try:
                        worksheet = workbook.worksheets[0]
                        with default_storage.open(out_rel, 'w') as csv_file:
                            writer = csv.writer(csv_file)
                            for row in worksheet.iter_rows(values_only=True):
                                writer.writerow(['' if value is None else value for value in row])
                    finally:
                        workbook.close()

        summary = {
            **(job.summary_json or {}),
            'format': export_format,
            'bytes': default_storage.size(out_rel),
            'source_storage_key': source_key,
        }
        SpreadsheetExportJob.objects.filter(id=job.id).update(
            status=SpreadsheetJobStatus.COMPLETED,
            output_storage_key=out_rel,
            summary_json=summary,
            finished_at=timezone.now(),
            error_text='',
        )
        log_action(
            organization_id=job.organization_id,
            actor_id=job.created_by_user_id,
            action=AuditLog.Action.EXPORT,
            entity_type='spreadsheet_export_job',
            entity_id=job.id,
            entity_label=job.document.title,
            diff={'status': SpreadsheetJobStatus.COMPLETED, 'output_storage_key': out_rel},
        )
    except Exception as exc:  # noqa: BLE001
        SpreadsheetExportJob.objects.filter(id=job.id).update(
            status=SpreadsheetJobStatus.FAILED,
            error_text=str(exc)[:4000],
            finished_at=timezone.now(),
        )
        logger.exception('process_export failed export_job_id=%s: %s', export_job_id, exc)
        raise
